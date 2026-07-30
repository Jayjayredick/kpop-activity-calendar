from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Notice
from .pipeline import notice_to_row


DAILY_COLUMNS = [
    "candidate_id","event_key","수집 채널","회사","레이블","아티스트","artist_id","활동 유형",
    "일정 판정","점수","매체 점수","게시일","제목","행사명","이벤트 날짜",
    "시작일","종료일","도시","공연장","클리핑 문구","매칭 키워드","출처 URL","기사 원문 URL",
    "네이버 뉴스 URL","매체","공식 소스","공식 확인","검색어","검증 상태",
    "검토 사유","제목 매칭 별칭","관련 기사 수","관련 기사 URL","기존 event_key",
    "source_id","notice_id","dedupe_key","수집시각",
]


def _frame(rows: list[dict], columns: list[str] | None = None) -> pd.DataFrame:
    if rows:
        return pd.DataFrame(rows)
    return pd.DataFrame(columns=columns or [])


def _style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for sheet in writer.book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 30
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            values = [str(cell.value or "") for cell in column[:100]]
            width = min(55, max(10, max((len(v) for v in values), default=10) + 2))
            if any(key in str(column[0].value) for key in ["URL", "클리핑", "제목", "행사명"]):
                width = min(55, max(width, 32))
            sheet.column_dimensions[letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_xlsx(
    path: str | Path,
    notices: list[Notice],
    excluded: list[dict],
    run_log: list[dict],
    params: dict,
    *,
    calendar_rows: list[dict] | None = None,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    selected_rows = [notice_to_row(n) for n in notices]
    raw_rows = [{**notice_to_row(n), "본문/검색 패시지": n.body} for n in notices]
    validation_rows = [
        row for row in selected_rows
        if row.get("검증 상태") == "REVIEW_REQUIRED"
    ]
    param_rows = [
        {
            "항목": key,
            "값": json.dumps(value, ensure_ascii=False)
            if isinstance(value, (dict, list))
            else value,
        }
        for key, value in params.items()
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _frame(selected_rows, DAILY_COLUMNS).to_excel(
            writer, sheet_name="daily_selected", index=False
        )
        _frame(calendar_rows or []).to_excel(
            writer, sheet_name="calendar_events", index=False
        )
        _frame(raw_rows, DAILY_COLUMNS + ["본문/검색 패시지"]).to_excel(
            writer, sheet_name="raw_articles", index=False
        )
        _frame(excluded).to_excel(writer, sheet_name="excluded", index=False)
        _frame(run_log).to_excel(writer, sheet_name="source_runs", index=False)
        _frame(validation_rows, DAILY_COLUMNS).to_excel(
            writer, sheet_name="validation_queue", index=False
        )
        _frame(param_rows, ["항목", "값"]).to_excel(
            writer, sheet_name="run_params", index=False
        )
        _style_workbook(writer)
    return path

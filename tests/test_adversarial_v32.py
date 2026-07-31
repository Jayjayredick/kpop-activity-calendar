from __future__ import annotations

import csv
import unittest
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from kpop_notice_collector.event_cluster import cluster_events
from kpop_notice_collector.event_parser import parse_event_fields
from kpop_notice_collector.exporter import export_xlsx
from kpop_notice_collector.models import Artist, Notice
from kpop_notice_collector.news_pipeline import run_naver_news
from kpop_notice_collector.review_store import upsert_review_queue
from kpop_notice_collector.schedule_compare import load_event_history


TZ = ZoneInfo("Asia/Seoul")
NOW = datetime(2026, 7, 31, 8, 0, tzinfo=TZ)


def make_notice(
    title: str,
    *,
    published_at: datetime = NOW,
    url: str = "https://news.example.com/article",
    activity_type: str = "COMEBACK",
    event_name: str = "",
) -> Notice:
    return Notice(
        source_id="naver_news",
        source_type="NAVER_NEWS",
        company="SM",
        label="SM Entertainment",
        artist_id="red_velvet",
        artist="Red Velvet",
        title=title,
        url=url,
        original_url=url,
        published_at=published_at,
        body="",
        fetched_at=NOW,
        activity_type=activity_type,
        event_name=event_name,
        matched_artist_alias="레드벨벳",
    )


class FakeAdapter:
    def __init__(self, notices: list[Notice]):
        self.notices = notices

    def collect_query(self, *_args, **_kwargs) -> list[Notice]:
        return list(self.notices)


def run_single_artist(notices: list[Notice]):
    artist = Artist(
        artist_id="red_velvet",
        company="SM",
        label="SM Entertainment",
        name="Red Velvet",
        aliases=["레드벨벳"],
        official_url="",
        source_ids=[],
    )
    config = {
        "hours": 24,
        "display": 100,
        "max_pages": 1,
        "min_score": 20,
        "blocked_title_aliases": [],
        "default_publisher_score": 3,
        "publisher_scores": {},
        "excluded_examples_per_reason": 20,
        "auto_publish_enabled": False,
    }
    return run_naver_news(
        [artist],
        config,
        history=[],
        now=NOW,
        adapter=FakeAdapter(notices),
    )


class TenAdversarialScenarios(unittest.TestCase):
    """v3.2 수정 후 새로 구성한 10개 상충·경계 상황."""

    def test_01_article_older_than_24_hours_is_excluded(self):
        old = make_notice(
            "레드벨벳, 8월 5일 새 앨범 발매 확정",
            published_at=NOW - timedelta(hours=30),
        )
        accepted, _excluded, log = run_single_artist([old])
        self.assertEqual(accepted, [])
        self.assertEqual(log[0]["outside_24h"], 1)

    def test_02_recent_article_is_collected_for_review(self):
        recent = make_notice(
            "레드벨벳, 8월 5일 새 앨범 발매 확정",
            published_at=NOW - timedelta(hours=2),
        )
        accepted, _excluded, _log = run_single_artist([recent])
        self.assertEqual(len(accepted), 1)
        self.assertEqual(accepted[0].event_start_date, "2026-08-05")
        self.assertEqual(accepted[0].validation_status, "REVIEW_REQUIRED")

    def test_03_tracking_parameters_do_not_create_duplicates(self):
        first = make_notice(
            "레드벨벳 컴백",
            url="https://news.example.com/view?article_id=111&utm_source=naver",
        )
        second = make_notice(
            "레드벨벳 컴백",
            url="https://news.example.com/view?utm_medium=feed&article_id=111",
        )
        self.assertEqual(first.dedupe_key, second.dedupe_key)

    def test_04_identity_query_parameters_remain_distinct(self):
        first = make_notice(
            "레드벨벳 컴백",
            url="https://news.example.com/view?article_id=111",
        )
        second = make_notice(
            "레드벨벳 컴백",
            url="https://news.example.com/view?article_id=222",
        )
        self.assertNotEqual(first.dedupe_key, second.dedupe_key)

    def test_05_duration_expression_is_not_calendar_date(self):
        notice = make_notice(
            "레드벨벳, 3일간 단독 콘서트 개최",
            activity_type="CONCERT",
        )
        parse_event_fields(notice)
        self.assertEqual(notice.event_start_date, "")
        self.assertEqual(notice.date_confidence, "NONE")

    def test_06_relative_delay_is_not_day_of_month(self):
        notice = make_notice(
            "레드벨벳, 3일 후 신보 발매 예정",
            activity_type="COMEBACK",
        )
        parse_event_fields(notice)
        self.assertEqual(notice.event_start_date, "")

    def test_07_month_only_schedule_stays_undated(self):
        notice = make_notice(
            "레드벨벳, 8월 서울 콘서트 개최",
            activity_type="CONCERT",
        )
        parse_event_fields(notice)
        self.assertEqual(notice.event_start_date, "")
        self.assertEqual(notice.date_confidence, "NONE")

    def test_08_two_competing_title_dates_are_conflict(self):
        notice = make_notice(
            "레드벨벳, 8월 10일·9월 12일 콘서트 개최",
            activity_type="CONCERT",
        )
        parse_event_fields(notice)
        self.assertTrue(notice.date_conflict)
        self.assertEqual(notice.date_confidence, "CONFLICT")
        self.assertEqual(notice.event_start_date, "")

    def test_09_same_date_and_city_but_different_events_stay_separate(self):
        common = {
            "published_at": NOW,
            "activity_type": "CONCERT",
        }
        first = make_notice(
            "레드벨벳 DREAM SHOW 8월 10일 서울 콘서트",
            url="https://news.example.com/dream",
            event_name="DREAM SHOW",
            **common,
        )
        second = make_notice(
            "레드벨벳 NEO CITY 8월 10일 서울 콘서트",
            url="https://news.example.com/neo",
            event_name="NEO CITY",
            **common,
        )
        for item in (first, second):
            item.event_start_date = "2026-08-10"
            item.event_end_date = "2026-08-10"
            item.event_dates = ["2026-08-10"]
            item.date_confidence = "HIGH"
            item.cities = ["서울"]
        self.assertEqual(len(cluster_events([first, second])), 2)

    def test_10_external_excel_formula_is_written_as_text(self):
        notice = make_notice(
            '=HYPERLINK("https://evil.example","click")',
            activity_type="COMEBACK",
            event_name="Safe Album",
        )
        with TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "tracker.xlsx"
            export_xlsx(target, [notice], [], [], {})
            workbook = load_workbook(target, data_only=False)
            sheet = workbook["daily_selected"]
            headers = {
                cell.value: cell.column
                for cell in sheet[1]
            }
            cell = sheet.cell(row=2, column=headers["제목"])
            self.assertEqual(cell.data_type, "s")
            self.assertTrue(str(cell.value).startswith("'="))


class AdditionalRegressionChecks(unittest.TestCase):
    def test_candidate_id_survives_event_name_parser_change(self):
        first = make_notice(
            "레드벨벳, 8월 10일 팬콘 개최",
            event_name="8월 10일 팬콘",
        )
        second = make_notice(
            "레드벨벳, 8월 10일 팬콘 개최",
            event_name="Red Velvet 팬콘서트",
        )
        self.assertEqual(first.candidate_id, second.candidate_id)

    def test_generic_artist_event_names_do_not_force_merge(self):
        first = make_notice(
            "레드벨벳 Birthday Party 개최",
            url="https://news.example.com/birthday",
            activity_type="CONCERT",
            event_name="Red Velvet 콘서트",
        )
        second = make_notice(
            "레드벨벳 Summer Night 개최",
            url="https://news.example.com/summer",
            activity_type="CONCERT",
            event_name="Red Velvet 콘서트",
        )
        for item in (first, second):
            item.event_start_date = "2026-08-10"
            item.event_end_date = "2026-08-10"
            item.event_dates = ["2026-08-10"]
            item.date_confidence = "HIGH"
            item.cities = ["서울"]
        self.assertEqual(len(cluster_events([first, second])), 2)

    def test_old_event_history_schema_is_not_used_for_schedule_compare(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "events_history.csv"
            path.write_text(
                "event_key,artist_id,activity_type,event_name,event_dates\n"
                "old,red_velvet,CONCERT,Old Show,2026-08-01\n",
                encoding="utf-8-sig",
            )
            self.assertEqual(load_event_history(path), [])

    def test_rejected_candidate_does_not_return_after_event_key_changes(self):
        notice = make_notice(
            "레드벨벳, 8월 10일 콘서트 개최",
            activity_type="CONCERT",
            event_name="Red Velvet Concert",
        )
        notice.event_start_date = "2026-08-10"
        notice.event_end_date = "2026-08-10"
        notice.event_dates = ["2026-08-10"]
        with TemporaryDirectory() as temp_dir:
            queue_path = Path(temp_dir) / "review_queue.csv"
            log_path = Path(temp_dir) / "review_log.csv"
            with log_path.open(
                "w", encoding="utf-8-sig", newline=""
            ) as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "reviewed_at",
                        "action",
                        "candidate_id",
                        "event_key",
                        "artist",
                        "activity_type",
                        "event_name",
                        "note",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "action": "REJECT",
                        "candidate_id": notice.candidate_id,
                        "event_key": "previous-parser-key",
                    }
                )
            rows = upsert_review_queue(
                queue_path,
                [notice],
                now=NOW,
                review_log_path=log_path,
            )
            self.assertEqual(rows, [])


if __name__ == "__main__":
    unittest.main()
